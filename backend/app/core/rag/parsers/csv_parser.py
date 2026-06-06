"""CSV / Excel parser — extracts structured data as readable text."""
import csv
from io import StringIO
from pathlib import Path
from app.core.rag.parsers.base import BaseParser


class CsvParser(BaseParser):
    """Parse CSV files into readable text.

    Converts each row to a key-value style representation,
    which embeds better for RAG retrieval than raw CSV.
    """

    supported_extensions = [".csv"]

    def parse(self, file_path: str) -> str:
        content = Path(file_path).read_text(encoding="utf-8-sig")
        reader = csv.DictReader(StringIO(content))
        rows = []
        for i, row in enumerate(reader):
            parts = [f"{k}: {v}" for k, v in row.items() if v]
            rows.append(f"Row {i + 1}: " + " | ".join(parts))
        return "\n".join(rows)


class ExcelParser(BaseParser):
    """Parse Excel files (.xlsx, .xls) into readable text.

    Reads all sheets, converts each to a labeled text block.
    """

    supported_extensions = [".xlsx", ".xls"]

    def parse(self, file_path: str) -> str:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    rows.append(row_text)
            if rows:
                parts.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
        wb.close()
        return "\n\n".join(parts)
